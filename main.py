import io
import json
import os
import re
import secrets
import urllib.parse
import uuid
from datetime import datetime
import httpx
from starlette.middleware.sessions import SessionMiddleware



def natural_key(project):
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', project.elp_project_id)]


def cell_to_html(value) -> str:
    """Convert an openpyxl cell value to HTML, preserving bold, italic, and colour."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    # CellRichText (openpyxl rich_text=True)
    try:
        parts = []
        for run in value:
            if isinstance(run, str):
                parts.append(run)
                continue
            text = run.text or ""
            font = getattr(run, "font", None)
            if font:
                color = getattr(font, "color", None)
                if color and getattr(color, "type", None) == "rgb":
                    rgb = color.rgb or ""
                    # rgb is 8-char ARGB; skip pure black and transparent
                    if len(rgb) == 8 and rgb.upper() not in ("FF000000", "00000000"):
                        text = f'<span style="color:#{rgb[2:]}">{text}</span>'
                if getattr(font, "italic", False):
                    text = f"<em>{text}</em>"
                if getattr(font, "bold", False):
                    text = f"<strong>{text}</strong>"
            parts.append(text)
        return "".join(parts)
    except Exception:
        return str(value)

import openpyxl
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.templating import Jinja2Templates
from sqlalchemy import or_
from sqlalchemy.orm import Session

from database import Base, engine, get_db
from models import Group, Preference, Project

from sqlalchemy import text
from sqlalchemy.orm import Session

Base.metadata.create_all(bind=engine)

# Database schema migration for groups table (adding email and token columns if missing)
db = Session(bind=engine)
try:
    columns_info = db.execute(text("PRAGMA table_info(groups)")).fetchall()
    column_names = [col[1] for col in columns_info]
    
    migrated = False
    if "student_1_email" not in column_names:
        print("Migrating database: Adding student email columns to 'groups' table...")
        db.execute(text("ALTER TABLE groups ADD COLUMN student_1_email VARCHAR"))
        db.execute(text("ALTER TABLE groups ADD COLUMN student_2_email VARCHAR"))
        db.execute(text("ALTER TABLE groups ADD COLUMN student_3_email VARCHAR"))
        db.execute(text("ALTER TABLE groups ADD COLUMN student_4_email VARCHAR"))
        db.execute(text("ALTER TABLE groups ADD COLUMN student_5_email VARCHAR"))
        db.commit()
        migrated = True
        
    if "token" not in column_names:
        print("Migrating database: Adding token column to 'groups' table...")
        db.execute(text("ALTER TABLE groups ADD COLUMN token VARCHAR"))
        db.commit()
        
        # Populate token for existing rows
        existing_groups = db.execute(text("SELECT id FROM groups WHERE token IS NULL")).fetchall()
        for g in existing_groups:
            db.execute(text("UPDATE groups SET token = :token WHERE id = :id"), {"token": str(uuid.uuid4()), "id": g[0]})
        db.commit()
        
        # Create unique index on the new token column
        try:
            db.execute(text("CREATE UNIQUE INDEX ix_groups_token ON groups (token)"))
            db.commit()
        except Exception as idx_err:
            print(f"Warning: Could not create unique index on token: {idx_err}")
            
        migrated = True
        
    # Database schema migration for projects table
    proj_columns_info = db.execute(text("PRAGMA table_info(projects)")).fetchall()
    proj_column_names = [col[1] for col in proj_columns_info]
    
    if "industry_type_1" not in proj_column_names:
        print("Migrating database: Adding new columns to 'projects' table...")
        db.execute(text("ALTER TABLE projects ADD COLUMN industry_type_1 VARCHAR"))
        db.execute(text("ALTER TABLE projects ADD COLUMN industry_type_2 VARCHAR"))
        db.execute(text("ALTER TABLE projects ADD COLUMN problem_category_1 VARCHAR"))
        db.execute(text("ALTER TABLE projects ADD COLUMN problem_category_2 VARCHAR"))
        db.execute(text("ALTER TABLE projects ADD COLUMN problem_category_3 VARCHAR"))
        db.commit()
        
        # Copy data from old columns if they exist
        if "industry_sector" in proj_column_names:
            db.execute(text("UPDATE projects SET industry_type_1 = industry_sector"))
        if "problem_type" in proj_column_names:
            db.execute(text("UPDATE projects SET problem_category_1 = problem_type"))
        db.commit()
        
        # Drop old columns
        if "industry_sector" in proj_column_names:
            try:
                db.execute(text("ALTER TABLE projects DROP COLUMN industry_sector"))
            except Exception as drop_err:
                print(f"Warning: Could not drop column industry_sector: {drop_err}")
        if "problem_type" in proj_column_names:
            try:
                db.execute(text("ALTER TABLE projects DROP COLUMN problem_type"))
            except Exception as drop_err:
                print(f"Warning: Could not drop column problem_type: {drop_err}")
        db.commit()
        migrated = True
        
    if migrated:
        print("Database migration completed successfully!")
except Exception as e:
    print(f"Error checking/running database migration: {e}")
finally:
    db.close()


app = FastAPI()

SESSION_SECRET = os.environ.get("SESSION_SECRET", "isb-elp-app-secure-session-key-2027")
app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET)

MICROSOFT_CLIENT_ID = os.environ.get("MICROSOFT_CLIENT_ID", "")
MICROSOFT_CLIENT_SECRET = os.environ.get("MICROSOFT_CLIENT_SECRET", "")
MICROSOFT_TENANT_ID = os.environ.get("MICROSOFT_TENANT_ID", "isb.edu")


def get_redirect_uri(request: Request) -> str:
    env_uri = os.environ.get("OAUTH_REDIRECT_URI", "")
    if env_uri:
        return env_uri
    proto = request.headers.get("x-forwarded-proto", "http")
    host = request.headers.get("host", "localhost:8000")
    return f"{proto}://{host}/auth/callback"

templates = Jinja2Templates(directory="templates")
security = HTTPBasic()

ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "elpadmin2027")


def require_admin(credentials: HTTPBasicCredentials = Depends(security)):
    valid = secrets.compare_digest(credentials.username, ADMIN_USERNAME) and \
            secrets.compare_digest(credentials.password, ADMIN_PASSWORD)
    if not valid:
        raise HTTPException(status_code=401, headers={"WWW-Authenticate": "Basic"})


# ---------------------------------------------------------------------------
# Student routes
# ---------------------------------------------------------------------------

def get_user_group(db: Session, user_email: str):
    if not user_email:
        return None
    return db.query(Group).filter(
        or_(
            Group.student_1_email == user_email,
            Group.student_2_email == user_email,
            Group.student_3_email == user_email,
            Group.student_4_email == user_email,
            Group.student_5_email == user_email,
        )
    ).first()


def get_projects_context(db: Session):
    projects = sorted(db.query(Project).all(), key=natural_key)
    industries = sorted({v for p in projects for v in [p.industry_type_1, p.industry_type_2] if v})
    problem_types = sorted({v for p in projects for v in [p.problem_category_1, p.problem_category_2, p.problem_category_3] if v})
    projects_json = json.dumps({
        p.elp_project_id: {
            "title": p.title,
            "industry_type_1": p.industry_type_1,
            "industry_type_2": p.industry_type_2 or "",
            "problem_category_1": p.problem_category_1,
            "problem_category_2": p.problem_category_2 or "",
            "problem_category_3": p.problem_category_3 or "",
            "problem_description": p.problem_description,
            "expected_outcomes": p.expected_outcomes,
        }
        for p in projects
    })
    return {
        "projects": projects,
        "industries": industries,
        "problem_types": problem_types,
        "projects_json": projects_json,
    }


@app.get("/", response_class=HTMLResponse)
def home(request: Request, db: Session = Depends(get_db)):
    user_email = request.session.get("user_email")
    if not user_email:
        return RedirectResponse(url="/landing-login", status_code=303)

    # Already registered (submitted or not) — go straight to their workspace.
    group = get_user_group(db, user_email)
    if group:
        return RedirectResponse(url=f"/submit/{group.token}", status_code=303)

    # Not registered yet — let them browse and shortlist projects freely.
    # Registering is a separate, optional step; their shortlist carries over
    # automatically once they do (see submit.html's localStorage migration).
    return templates.TemplateResponse("submit.html", {
        "request": request,
        "group": None,
        "token": None,
        **get_projects_context(db),
        "user_email": user_email,
        "user_name": request.session.get("user_name", ""),
        "user_token": None,
        "user_group_id": None,
    })


@app.get("/landing-login", response_class=HTMLResponse)
def landing_login(request: Request):
    user_email = request.session.get("user_email")
    if user_email:
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse("login.html", {
        "request": request,
        "dev_mode": not bool(MICROSOFT_CLIENT_ID),
    })


@app.get("/login")
def login(request: Request):
    if not MICROSOFT_CLIENT_ID:
        raise HTTPException(status_code=500, detail="MICROSOFT_CLIENT_ID environment variable is not set.")
    
    redirect_uri = get_redirect_uri(request)
    state = secrets.token_urlsafe(16)
    request.session["oauth_state"] = state
    
    microsoft_auth_url = (
        f"https://login.microsoftonline.com/{MICROSOFT_TENANT_ID}/oauth2/v2.0/authorize"
        "?response_type=code"
        f"&client_id={MICROSOFT_CLIENT_ID}"
        f"&redirect_uri={urllib.parse.quote(redirect_uri)}"
        "&response_mode=query"
        "&scope=openid%20profile%20email%20User.Read"
        f"&state={state}"
        "&prompt=select_account"
    )
    return RedirectResponse(url=microsoft_auth_url)


@app.get("/auth/callback")
async def auth_callback(request: Request, code: str = None, state: str = None, error: str = None, db: Session = Depends(get_db)):
    if error:
        error_desc = request.query_params.get("error_description", error)
        return templates.TemplateResponse("login.html", {"request": request, "error": f"Microsoft Sign-In failed: {error_desc}"}, status_code=400)
    
    if not code or not state:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Invalid auth request. Missing code or state."}, status_code=400)
    
    # Verify state
    saved_state = request.session.pop("oauth_state", None)
    if not saved_state or saved_state != state:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Session expired or state mismatch. Please try again."}, status_code=400)
    
    redirect_uri = get_redirect_uri(request)
    
    # Exchange code for tokens
    async with httpx.AsyncClient() as client:
        try:
            token_response = await client.post(
                f"https://login.microsoftonline.com/{MICROSOFT_TENANT_ID}/oauth2/v2.0/token",
                data={
                    "code": code,
                    "client_id": MICROSOFT_CLIENT_ID,
                    "client_secret": MICROSOFT_CLIENT_SECRET,
                    "redirect_uri": redirect_uri,
                    "grant_type": "authorization_code",
                },
                timeout=10.0,
            )
            token_data = token_response.json()
            if "error" in token_data:
                return templates.TemplateResponse("login.html", {"request": request, "error": f"Token exchange failed: {token_data.get('error_description', token_data['error'])}"}, status_code=400)
            
            access_token = token_data.get("access_token")
            if not access_token:
                return templates.TemplateResponse("login.html", {"request": request, "error": "Failed to retrieve access token from Microsoft."}, status_code=400)
            
            # Fetch user info from Microsoft Graph
            userinfo_response = await client.get(
                "https://graph.microsoft.com/v1.0/me",
                headers={"Authorization": f"Bearer {access_token}"},
                timeout=10.0,
            )
            user_info = userinfo_response.json()
        except Exception as e:
            return templates.TemplateResponse("login.html", {"request": request, "error": f"Network error during authentication: {e}"}, status_code=500)
            
    # Resolve email (mail is preferred, userPrincipalName is the login fallback)
    email = user_info.get("mail") or user_info.get("userPrincipalName", "")
    email = email.strip().lower()
    name = user_info.get("displayName", "").strip()
    
    if not email:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Microsoft profile did not return a valid email address."}, status_code=400)
    
    # Verify domain restriction: Only members from isb.edu are allowed.
    allowed_domains = ["isb.edu"]
    is_allowed = False
    for domain in allowed_domains:
        if email.endswith(f"@{domain}") or email.endswith(f".{domain}"):
            is_allowed = True
            break
            
    # Dev bypasses for testing
    if email.endswith("@pg.fju.us") or "praveengarimella" in email or "vishal" in email:
        is_allowed = True
        
    if not is_allowed:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": f"Access restricted. Email '{email}' does not belong to the @isb.edu domain."
        }, status_code=403)
        
    # User is authorized! Set session
    request.session["user_email"] = email
    request.session["user_name"] = name
    request.session["user_picture"] = ""
    
    # Check if the user is already part of a group
    group = db.query(Group).filter(
        or_(
            Group.student_1_email == email,
            Group.student_2_email == email,
            Group.student_3_email == email,
            Group.student_4_email == email,
            Group.student_5_email == email,
        )
    ).first()
    
    if group:
        return RedirectResponse(url=f"/submit/{group.token}", status_code=303)
    else:
        return RedirectResponse(url="/register", status_code=303)



@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/", status_code=303)


# ---------------------------------------------------------------------------
# Local-only dev login — bypasses Microsoft OAuth.
# Disabled whenever MICROSOFT_CLIENT_ID is set, so it never activates in
# production / any environment configured with real Azure AD credentials.
# ---------------------------------------------------------------------------

@app.get("/dev-login", response_class=HTMLResponse)
def dev_login_form(request: Request):
    if MICROSOFT_CLIENT_ID:
        raise HTTPException(status_code=404)
    return templates.TemplateResponse("dev_login.html", {"request": request})


@app.post("/dev-login")
def dev_login_submit(request: Request, email: str = Form(...), name: str = Form(...)):
    if MICROSOFT_CLIENT_ID:
        raise HTTPException(status_code=404)
    request.session["user_email"] = email.strip().lower()
    request.session["user_name"] = name.strip()
    request.session["user_picture"] = ""
    return RedirectResponse(url="/", status_code=303)


@app.get("/register", response_class=HTMLResponse)
def register_form(request: Request, db: Session = Depends(get_db)):
    user_email = request.session.get("user_email")
    if not user_email:
        return RedirectResponse(url="/landing-login", status_code=303)
        
    # Check if user is already in a group
    group = db.query(Group).filter(
        or_(
            Group.student_1_email == user_email,
            Group.student_2_email == user_email,
            Group.student_3_email == user_email,
            Group.student_4_email == user_email,
            Group.student_5_email == user_email,
        )
    ).first()
    if group:
        return RedirectResponse(url=f"/submit/{group.token}", status_code=303)

    return templates.TemplateResponse("register.html", {
        "request": request,
        "user_email": user_email,
        "user_name": request.session.get("user_name", "")
    })


@app.post("/register")
async def register_group(
    request: Request,
    db: Session = Depends(get_db),
    group_id_suffix: str = Form(...),
    s1_name: str = Form(...), s1_roll: str = Form(...),
    s2_name: str = Form(...), s2_roll: str = Form(...), s2_email: str = Form(...),
    s3_name: str = Form(...), s3_roll: str = Form(...), s3_email: str = Form(...),
    s4_name: str = Form(...), s4_roll: str = Form(...), s4_email: str = Form(...),
    s5_name: str = Form(...), s5_roll: str = Form(...), s5_email: str = Form(...),
):
    user_email = request.session.get("user_email")
    if not user_email:
        return RedirectResponse(url="/landing-login", status_code=303)
        
    group_id = "ELP" + group_id_suffix.strip().upper()
    s1_email = user_email
    
    s2_email = s2_email.strip().lower()
    s3_email = s3_email.strip().lower()
    s4_email = s4_email.strip().lower()
    s5_email = s5_email.strip().lower()
    
    emails = [s1_email, s2_email, s3_email, s4_email, s5_email]
    rolls = [r.strip() for r in [s1_roll, s2_roll, s3_roll, s4_roll, s5_roll]]

    def render_register(error=None, already_exists=False, existing_token=None):
        return templates.TemplateResponse("register.html", {
            "request": request,
            "error": error,
            "already_exists": already_exists,
            "existing_token": existing_token,
            "user_email": user_email,
            "user_name": request.session.get("user_name", ""),
            "form_data": {
                "group_id_suffix": group_id_suffix,
                "s1_name": s1_name, "s1_roll": s1_roll,
                "s2_name": s2_name, "s2_roll": s2_roll, "s2_email": s2_email,
                "s3_name": s3_name, "s3_roll": s3_roll, "s3_email": s3_email,
                "s4_name": s4_name, "s4_roll": s4_roll, "s4_email": s4_email,
                "s5_name": s5_name, "s5_roll": s5_roll, "s5_email": s5_email,
            },
        }, status_code=400)

    # 1. Enforce that all email domains are isb.edu (with dev bypass)
    for email in emails:
        is_allowed = False
        if email.endswith("@isb.edu") or email.endswith(".isb.edu") or email.endswith("@pg.fju.us") or "praveengarimella" in email or "vishal" in email:
            is_allowed = True
        if not is_allowed:
            return render_register(error=f"All group member emails must belong to the @isb.edu domain. Checked: {email}")

    # 2. Check for duplicate emails within the form
    if len(set(emails)) != 5:
        return render_register(error="Duplicate student emails detected. Each group member must have a unique email.")

    # 2b. Check for duplicate roll numbers within the form
    if len(set(rolls)) != 5:
        return render_register(error="Each student must have a unique roll number. Please check your entries.")

    # 3. Same group ID already registered
    existing = db.query(Group).filter(Group.group_id == group_id).first()
    if existing:
        if existing.is_submitted:
            return render_register(
                error=f"Group {group_id} has already submitted preferences. You can view your submission at the link below.",
                already_exists=True,
                existing_token=existing.token
            )
        else:
            if user_email in [existing.student_1_email, existing.student_2_email, existing.student_3_email, existing.student_4_email, existing.student_5_email]:
                return RedirectResponse(url=f"/submit/{existing.token}", status_code=303)
            else:
                return render_register(error=f"Group {group_id} is already registered by another representative.")

    # 4. Check if any member (roll OR email) is already registered in ANY group
    for email in emails:
        already_in_group = db.query(Group).filter(
            or_(
                Group.student_1_email == email,
                Group.student_2_email == email,
                Group.student_3_email == email,
                Group.student_4_email == email,
                Group.student_5_email == email,
            )
        ).first()
        if already_in_group:
            return render_register(error=f"Student email {email} is already registered in group {already_in_group.group_id}.")

    for r in rolls:
        already_in_group = db.query(Group).filter(
            or_(
                Group.student_1_roll == r,
                Group.student_2_roll == r,
                Group.student_3_roll == r,
                Group.student_4_roll == r,
                Group.student_5_roll == r,
            )
        ).first()
        if already_in_group:
            return render_register(error=f"Student roll number {r} is already registered in group {already_in_group.group_id}.")

    group = Group(
        group_id=group_id,
        token=str(uuid.uuid4()),
        student_1_name=s1_name.strip(), student_1_roll=rolls[0], student_1_email=s1_email,
        student_2_name=s2_name.strip(), student_2_roll=rolls[1], student_2_email=s2_email,
        student_3_name=s3_name.strip(), student_3_roll=rolls[2], student_3_email=s3_email,
        student_4_name=s4_name.strip(), student_4_roll=rolls[3], student_4_email=s4_email,
        student_5_name=s5_name.strip(), student_5_roll=rolls[4], student_5_email=s5_email,
        is_submitted=False,
    )
    db.add(group)
    db.commit()
    db.refresh(group)
    return RedirectResponse(url=f"/submit/{group.token}", status_code=303)


@app.get("/submit/{token}", response_class=HTMLResponse)
def submit_page(token: str, request: Request, db: Session = Depends(get_db)):
    user_email = request.session.get("user_email")
    if not user_email:
        return RedirectResponse(url="/landing-login", status_code=303)
    group = db.query(Group).filter(Group.token == token).first()
    if not group:
        return templates.TemplateResponse("not_found.html", {"request": request}, status_code=404)

    # Secure authorization check
    member_emails = [group.student_1_email, group.student_2_email, group.student_3_email, group.student_4_email, group.student_5_email]
    if user_email not in member_emails:
        return HTMLResponse("<h1>403 Forbidden</h1><p>You do not have permission to access this group's workspace.</p>", status_code=403)

    if group.is_submitted:
        prefs = sorted(group.preferences, key=lambda x: x.rank)
        return templates.TemplateResponse("submitted.html", {
            "request": request,
            "group": group,
            "token": token,
            "prefs": prefs,
            "user_email": user_email,
            "user_name": request.session.get("user_name", ""),
            "user_token": token,
            "user_group_id": group.group_id,
        })

    return templates.TemplateResponse("submit.html", {
        "request": request,
        "group": group,
        "token": token,
        **get_projects_context(db),
        "user_email": user_email,
        "user_name": request.session.get("user_name", ""),
        "user_token": token,
        "user_group_id": group.group_id,
    })


@app.post("/submit/{token}")
async def submit_preferences(
    token: str,
    request: Request,
    db: Session = Depends(get_db),
    pref_1: str = Form(...), pref_2: str = Form(...),
    pref_3: str = Form(...), pref_4: str = Form(...),
    pref_5: str = Form(...), pref_6: str = Form(...),
    pref_7: str = Form(...), pref_8: str = Form(...),
    pref_9: str = Form(...), pref_10: str = Form(...),
):
    user_email = request.session.get("user_email")
    if not user_email:
        return RedirectResponse(url="/landing-login", status_code=303)
    group = db.query(Group).filter(Group.token == token).first()
    if not group:
        return templates.TemplateResponse("not_found.html", {"request": request}, status_code=404)
        
    # Secure authorization check
    member_emails = [group.student_1_email, group.student_2_email, group.student_3_email, group.student_4_email, group.student_5_email]
    if user_email not in member_emails:
        return HTMLResponse("<h1>403 Forbidden</h1><p>You do not have permission to modify this group's preferences.</p>", status_code=403)

    if group.is_submitted:
        return RedirectResponse(url=f"/submit/{token}", status_code=303)

    prefs = [pref_1, pref_2, pref_3, pref_4, pref_5,
             pref_6, pref_7, pref_8, pref_9, pref_10]

    if len(set(prefs)) != 10:
        return templates.TemplateResponse("submit.html", {
            "request": request,
            "group": group,
            "token": token,
            **get_projects_context(db),
            "error": "Duplicate projects detected. Please ensure all 10 preferences are different.",
            "user_email": user_email,
            "user_name": request.session.get("user_name", ""),
            "user_token": token,
            "user_group_id": group.group_id,
        }, status_code=400)

    for rank, pid in enumerate(prefs, start=1):
        db.add(Preference(group_id=group.group_id, elp_project_id=pid, rank=rank))

    group.is_submitted = True
    group.submitted_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(url=f"/submit/{token}", status_code=303)


@app.get("/submit/{token}/download")
def download_preferences(token: str, request: Request, db: Session = Depends(get_db)):
    user_email = request.session.get("user_email")
    if not user_email:
        raise HTTPException(status_code=401, detail="Unauthorized")
    group = db.query(Group).filter(Group.token == token).first()
    if not group or not group.is_submitted:
        raise HTTPException(status_code=404)

    # Secure authorization check
    member_emails = [group.student_1_email, group.student_2_email, group.student_3_email, group.student_4_email, group.student_5_email]
    if user_email not in member_emails:
        raise HTTPException(status_code=403, detail="Forbidden")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Preferences"

    # Group info header rows
    ws.append(["Group ID", group.group_id])
    ws.append(["Submitted At", group.submitted_at.strftime("%d %b %Y, %H:%M")])
    ws.append([])
    ws.append(["#", "Student Name", "Roll Number", "Email Address"])
    for i, (name, roll, email) in enumerate([
        (group.student_1_name, group.student_1_roll, group.student_1_email),
        (group.student_2_name, group.student_2_roll, group.student_2_email),
        (group.student_3_name, group.student_3_roll, group.student_3_email),
        (group.student_4_name, group.student_4_roll, group.student_4_email),
        (group.student_5_name, group.student_5_roll, group.student_5_email),
    ], start=1):
        ws.append([i, name, roll, email])

    ws.append([])
    ws.append(["Serial Number", "Project ID", "Title", "Problem Type", "Industry Sector"])
    for pref in sorted(group.preferences, key=lambda x: x.rank):
        ws.append([
            pref.rank,
            pref.elp_project_id,
            pref.project.title,
            pref.project.problem_category_1,
            pref.project.industry_type_1,
        ])

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"ELP_preferences_{group.group_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )



# ---------------------------------------------------------------------------
# Admin routes
# ---------------------------------------------------------------------------

@app.get("/admin", response_class=HTMLResponse)
def admin_dashboard(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    return templates.TemplateResponse("admin/dashboard.html", {
        "request": request,
        "project_count": db.query(Project).count(),
        "group_count": db.query(Group).count(),
        "submitted_count": db.query(Group).filter(Group.is_submitted == True).count(),
        "recent_groups": db.query(Group).order_by(Group.submitted_at.desc()).limit(10).all(),
    })


@app.post("/admin/upload", response_class=HTMLResponse)
async def upload_projects(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    def dashboard(error=None, success=None):
        return templates.TemplateResponse("admin/dashboard.html", {
            "request": request,
            "project_count": db.query(Project).count(),
            "group_count": db.query(Group).count(),
            "submitted_count": db.query(Group).filter(Group.is_submitted == True).count(),
            "recent_groups": db.query(Group).order_by(Group.submitted_at.desc()).limit(10).all(),
            "error": error,
            "success": success,
        })

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), rich_text=True)
    except Exception:
        return dashboard(error="Could not read the file. Please upload a valid .xlsx file.")

    ws = wb.active
    raw_headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    headers = [h.lower().replace(" ", "_") for h in raw_headers]

    col = {}
    for i, h in enumerate(headers):
        if "project_id" in h or "project_code" in h or (h.startswith("elp") and "id" in h):
            col["elp_project_id"] = i
        elif h == "title" or "project_title" in h:
            col["title"] = i
        elif "industry" in h and "2" in h:
            col["industry_type_2"] = i
        elif "industry" in h:
            col["industry_type_1"] = i
        elif ("problem_type" in h or "category" in h) and "3" in h:
            col["problem_category_3"] = i
        elif ("problem_type" in h or "category" in h) and "2" in h:
            col["problem_category_2"] = i
        elif "problem_type" in h or "category" in h:
            col["problem_category_1"] = i
        elif "description" in h:
            col["problem_description"] = i
        elif "outcome" in h:
            col["expected_outcomes"] = i

    missing = [k for k in ("elp_project_id", "title", "industry_type_1", "problem_category_1",
                            "problem_description", "expected_outcomes") if k not in col]
    if missing:
        readable = [m.replace("_", " ").title() for m in missing]
        return dashboard(error=f"Could not find these columns: {', '.join(readable)}. "
                               f"Found: {', '.join(raw_headers)}")

    def get_cell_str(row, key):
        if key not in col or col[key] >= len(row):
            return ""
        return str(row[col[key]].value or "").strip()

    def get_cell_val(row, key):
        if key not in col or col[key] >= len(row):
            return None
        return row[col[key]].value

    uploaded = updated = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        # Skip completely empty rows
        if not any(cell.value is not None for cell in row):
            continue
            
        # Guard: Check if the row has the required columns index bounds
        max_idx = max(col.values())
        if len(row) <= max_idx:
            return dashboard(error=f"Row {row_idx} is incomplete or has missing columns. Please check your data.")

        pid = str(row[col["elp_project_id"]].value or "").strip()
        if not pid:
            continue
        data = dict(
            title=cell_to_html(get_cell_val(row, "title")).strip(),
            industry_type_1=get_cell_str(row, "industry_type_1"),
            industry_type_2=get_cell_str(row, "industry_type_2"),
            problem_category_1=get_cell_str(row, "problem_category_1"),
            problem_category_2=get_cell_str(row, "problem_category_2"),
            problem_category_3=get_cell_str(row, "problem_category_3"),
            problem_description=cell_to_html(get_cell_val(row, "problem_description")),
            expected_outcomes=cell_to_html(get_cell_val(row, "expected_outcomes")),
        )
        existing = db.query(Project).filter(Project.elp_project_id == pid).first()
        if existing:
            for k, v in data.items():
                setattr(existing, k, v)
            updated += 1
        else:
            db.add(Project(elp_project_id=pid, **data))
            uploaded += 1

    db.commit()
    return dashboard(success=f"Done — {uploaded} new projects added, {updated} updated.")


@app.get("/admin/export")
def export_preferences(db: Session = Depends(get_db), _=Depends(require_admin)):
    wb = openpyxl.Workbook()

    # Sheet 1 — full detail
    ws1 = wb.active
    ws1.title = "Preferences"
    ws1.append([
        "Group ID",
        "S1 Name", "S1 Roll", "S1 Email",
        "S2 Name", "S2 Roll", "S2 Email",
        "S3 Name", "S3 Roll", "S3 Email",
        "S4 Name", "S4 Roll", "S4 Email",
        "S5 Name", "S5 Roll", "S5 Email",
        "Pref 1", "Pref 2", "Pref 3", "Pref 4", "Pref 5",
        "Pref 6", "Pref 7", "Pref 8", "Pref 9", "Pref 10",
        "Submitted At",
    ])

    groups = db.query(Group).filter(Group.is_submitted == True).order_by(Group.submitted_at).all()
    for group in groups:
        pref_ids = [p.elp_project_id for p in sorted(group.preferences, key=lambda x: x.rank)]
        while len(pref_ids) < 10:
            pref_ids.append("")
        ws1.append([
            group.group_id,
            group.student_1_name, group.student_1_roll, group.student_1_email,
            group.student_2_name, group.student_2_roll, group.student_2_email,
            group.student_3_name, group.student_3_roll, group.student_3_email,
            group.student_4_name, group.student_4_roll, group.student_4_email,
            group.student_5_name, group.student_5_roll, group.student_5_email,
            *pref_ids,
            group.submitted_at.strftime("%Y-%m-%d %H:%M:%S"),
        ])


    # Sheet 2 — group / project pairs for matching algorithm
    ws2 = wb.create_sheet(title="Group-Project Pairs")
    ws2.append(["Group ID", "Preferred Project ID"])
    for group in groups:
        for pref in sorted(group.preferences, key=lambda x: x.rank):
            ws2.append([group.group_id, pref.elp_project_id])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=elp_preferences.xlsx"},
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
